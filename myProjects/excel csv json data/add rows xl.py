#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.
# This program used for finding and updating corresponding information in excel,, eg adding rows or colums after a specific data type
import openpyxl, os

os.chdir('C:\\Users\\dhira\\Desktop\\mywork')


'''SAMPLE DATA
PRODUCE	COST PER POUND	POUNDS SOLD	TOTAL
Potatoes	0.86	21.6	18.58
Okra	2.26	38.6	87.24
Fava beans	2.69	32.8	88.23
Watermelon	0.66	27.3	18.02
Garlic	1.19	4.9	5.83
Parsnips	2.27	1.1	2.5
Asparagus	2.49	37.9	94.37
Avocados	3.23	9.2	29.72
Celery	3.07	28.9	88.72
Okra	2.26	40	90.4
Spinach	4.12	30	123.6
Cucumber	1.07	36	38.52
Apricots	3.71	29.4	109.07
Okra	2.26	9.5	21.47
Fava beans	2.69	5.3	14.26
Watermelon	0.66	35.4	23.36
Ginger	5.13	14.4	73.87
Corn	1.07	12.2	13.05
Grapefruit	0.76	35.7	27.13
Ginger	5.13	15.2	77.98
'''




wb = openpyxl.load_workbook('example2.xlsx')
sheet = wb['Sheet1']

# The produce types and their corresponding number of rows to add
ADD_ROWS = {'Garlic': 2,
            'Celery': 4,
            'Lemon': 1}

 # Loop through the rows and add rows
for rowNum in range(2, sheet.max_row):  # skip the first row
     produceName = sheet.cell(row=rowNum, column=1).value
     if produceName in ADD_ROWS:
         print(rowNum)
         sheet.insert_rows(rowNum+1, ADD_ROWS[produceName])
         
wb.save('example3.xlsx')



'''PRODUCE	COST PER POUND	POUNDS SOLD	TOTAL
Potatoes	0.86	21.6	18.58
Okra	2.26	38.6	87.24
Fava beans	2.69	32.8	88.23
Watermelon	0.66	27.3	18.02
Garlic	1.19	4.9	5.83
			
			
Parsnips	2.27	1.1	0
Asparagus	2.49	37.9	0
Avocados	3.23	9.2	2.5
Celery	3.07	28.9	94.37
			
			
			
			
Okra	2.26	40	29.72
Spinach	4.12	30	88.72'''

