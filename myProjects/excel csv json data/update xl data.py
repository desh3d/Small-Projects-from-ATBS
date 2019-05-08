#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.
# This program used for finding and updating corresponding information in excel
import openpyxl, os

os.chdir('C:\\Users\\dhira\\Desktop\\mywork')

'''SAMPLE DATA  PRODUCE	COST PER POUND	POUNDS SOLD	TOTAL
Potatoes	0.86	21.6	18.58
Okra	2.26	38.6	87.24
Fava beans	2.69	32.8	88.23
Watermelon	0.66	27.3	18.02
Garlic	1.19	4.9	5.83
Parsnips	2.27	1.1	2.5
Asparagus	2.49	37.9	94.37
Avocados	3.23	9.2	29.72
Celery	3.07	28.9	88.72
Okra	2.26	40	90.4'''




wb = openpyxl.load_workbook('example2.xlsx')
sheet = wb['Sheet1']

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

 # Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):  # skip the first row
     produceName = sheet.cell(row=rowNum, column=1).value
     if produceName in PRICE_UPDATES:
           sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
wb.save('updatedProduceSales.xlsx')
